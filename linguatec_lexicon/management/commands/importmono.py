import json
import os

from django.core.management import BaseCommand, CommandError
from django.db import transaction
from odf import opendocument
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import load_workbook

from linguatec_lexicon.models import Entry, Lexicon, Word


def is_row_empty(row):
    for cell in row:
        if cell.value is not None:  # If any cell in the row has a value
            return False
    return True


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument(
            'lexicon_code', type=str,
            help="Select the lexicon where data will be imported",
        )
        parser.add_argument('input_file', type=str)

    @transaction.atomic
    def handle(self, *args, **options):
        self.input_file = options['input_file']
        self.lexicon_code = options['lexicon_code']

        # check that a lexicon with that code exist
        try:
            self.lexicon = Lexicon.objects.get_by_slug(self.lexicon_code)
        except Lexicon.DoesNotExist:
            raise CommandError('Error: There is not a lexicon with that code: ' + self.lexicon_code)

        # self.xlsx = pd.read_excel(self.input_file, sheet_name=None, header=None, usecols="A:E",
        #                           names=["term", "url", "etimol", "def", "def2"])

        self._words = set()

        # use temp lexicon and if everything is ok, then change to the real one
        self._lexicon = self.lexicon
        self.lexicon = Lexicon.objects.create(name="Temp Lexicon", src_language="99", dst_language="99")

        self.xlsx = load_workbook(self.input_file, read_only=True)
        self.ods = self.xlsx2ods()
        sheet = self.xlsx.active

        errors = []
        for i, row in enumerate(sheet.iter_rows()):
            if i == 0:
                continue

            row_number = i + 1
            if is_row_empty(row):
                print(f"Row {row_number} is empty")
                continue

            wrow = self.validate_row(row, row_number)
            if wrow.errors:
                errors.append({
                    "row": row_number,
                    "term": wrow.term,
                    "errors": wrow.errors,
                })
                continue

            word = Word(
                lexicon=self.lexicon,
                term=wrow.term,
                etimol=wrow.etimol,
            )
            word.save()
            entries = [Entry(word=word, translation=wrow.definition)]
            if wrow.definition2:
                entries.append(Entry(word=word, translation=wrow.definition2))
            word.entries.bulk_create(entries)

            # TODO: think how to handle too many errors
            if len(errors) >= 100:
                break

        if errors:
            self.lexicon.delete()
            self.print_errors(errors, format="json")
        else:
            self.lexicon.words.update(lexicon=self._lexicon)
            self.lexicon.delete()
            self.lexicon = self._lexicon

    def xlsx2ods(self):
        # TODO(@slamora): create code to convert xlsx to ods
        odspath = self.input_file.replace(".xlsx", ".ods")
        if not os.path.exists(odspath):
            raise CommandError(f"ODS file doesn't exist: {odspath}")
        return opendocument.load(odspath)

    def validate_row(self, row, row_number):
        instance = Row(row, row_number, self._words)
        instance.is_valid()

        instance.etimol = self.extract_etimol_rich_text(row_number)

        return instance

    def extract_etimol_rich_text(self, row_number):
        # retrieve etimol XML from ODS
        doc = self.ods
        table = doc.getElementsByType(Table)[0]
        row = table.getElementsByType(TableRow)[row_number - 1]  # 0-indexed
        cell = row.getElementsByType(TableCell)[2]

        etimol_html = extract_text_as_html(cell)

        return etimol_html

    def print_errors(self, errors, format="json"):
        for error in errors:
            if format == "json":
                for key, value in error["errors"].items():
                    self.stdout.write(self.style.ERROR(
                        json.dumps({
                            "word": f"#{error['row']}: {error['term']}",
                            "column": key,
                            "message": value,
                        })
                    ))

                continue

            # extended output
            self.stdout.write(self.style.ERROR(f"Row: {error['row']}"))
            self.stdout.write(self.style.ERROR(f"Term: {error['term']}"))
            for key, value in error["errors"].items():
                self.stdout.write(self.style.ERROR(f"  * {key}: {value}"))
            self.stdout.write("\n")


class Row:
    TERM = 0
    URL = 1
    ETIMOL = 2
    DEFINITION = 4  # legacy 3
    DEFINITION2 = 6  # legacy 4

    def __init__(self, row, row_number, words):
        self.term = row[self.TERM].value
        self.url = row[self.URL].value
        self.etimol = row[self.ETIMOL].value or ""
        self.definition = row[self.DEFINITION].value
        self.definition2 = row[self.DEFINITION2].value
        self.row_number = row_number

        self.existing_words = words

    def is_valid(self):
        self.errors = {}
        if not self.term:
            self.errors["term"] = "Term is required"
        if not self.definition:
            self.errors["definition"] = "Definition is required"

        self.validate_unique()

        return False if self.errors else True

    def validate_unique(self):
        if self.term in self.existing_words:
            self.errors["term"] = "This term already exists"
            return False

        # TODO(@slamora): check if the term is in the database???? or all the imports are from scratch???
        # allow user to decide if the import is from scratch or not
        # TODO(@slamora): optimize storing on memmory to perform only one query
        # if Word.objects.filter(lexicon=XXX, term=self.term).exists():
        #     self.errors["term"] = "This term already exists"

        self.existing_words.add(self.term)
        return True


def extract_text_as_html(cell):
    html_content = ""
    paragraphs = cell.getElementsByType(P)
    for p in paragraphs:
        for element in p.childNodes:

            if element.tagName == 'text:span':  # Handle styled spans
                text = element.firstChild.data if element.firstChild else ""

                if element.getAttribute("stylename") == "T1":   # T1 is the style for italic text
                    html_content += f"<i>{text}</i>"
                else:
                    html_content += text

            elif element.tagName == 'text:s':  # Handle spaces
                html_content += " "
            else:  # Handle plain text
                text = element.data if element.nodeType == element.TEXT_NODE else ""
                html_content += text
    return html_content
